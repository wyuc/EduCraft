import os
import json
import argparse
import glob
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from storage import ScriptStorage
from utils import get_images_dir
from config import BASE_DIR

def natural_sort_key(s):
    """
    Natural sort key function for sorting filenames with numbers.
    This ensures that e.g., "2.png" comes before "10.png"
    """
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', s)]

def scripts_to_excel(ppt_path, algo, model_provider, output_dir=None):
    """
    Convert lecture scripts and images to Excel using the SQL storage.
    
    Args:
        ppt_path (str): Path to the PowerPoint file
        algo (str): Algorithm used for generation (e.g., 'vlm')
        model_provider (str): Model provider name (e.g., 'claude', 'gemini')
        output_dir (str, optional): Directory to save the Excel file. If None, uses the PPT directory
    """
    # Get storage instance
    storage = ScriptStorage()
    
    # Get the latest completed task matching the criteria
    latest_task = storage.get_latest_completed_task(ppt_path, algo, model_provider)
    
    if not latest_task:
        print(f"Error: No completed tasks found for {ppt_path} with algorithm {algo} and provider {model_provider}")
        return
    
    task_id = latest_task['task_id']
    
    # Get all slides for this task
    slides = storage.get_all_slides(task_id)
    if not slides:
        print(f"Error: No slides found for task {task_id}")
        return
    
    images_dir = get_images_dir(ppt_path)
    
    # Check if images directory exists
    if not os.path.exists(images_dir):
        print(f"Warning: Images directory not found at {images_dir}")
        images_dir = None
    
    # Find image files if available
    image_files = []
    if images_dir:
        image_files = glob.glob(os.path.join(images_dir, '*.*'))
        image_extensions = ['.png', '.jpg', '.jpeg', '.gif', '.bmp']
        image_files = [f for f in image_files if os.path.splitext(f)[1].lower() in image_extensions]
        
        # Use natural sorting to ensure correct numerical order
        image_files.sort(key=natural_sort_key)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Lecture"
    
    # Set column headers
    ws['A1'] = "Slide"
    ws['B1'] = "Script"
    
    # Set column widths
    ws.column_dimensions['A'].width = 70
    ws.column_dimensions['B'].width = 80
    
    # Convert slides to dict for easier access by slide number
    slides_dict = {slide['slide_num']: slide['content'] for slide in slides}
    
    # Add content
    max_slides = max(len(image_files), len(slides)) if image_files else len(slides)
    
    for i in range(max_slides):
        row = i + 2  # Start from row 2 (after headers)
        
        # Add image if available
        if image_files and i < len(image_files):
            try:
                # Get image dimensions
                img = PILImage.open(image_files[i])
                width, height = img.size
                
                # Resize image for Excel
                max_height = 400
                max_width = 450
                scale = min(max_width/width, max_height/height)
                
                img_excel = Image(image_files[i])
                img_excel.width = width * scale
                img_excel.height = height * scale
                
                # Add to cell
                ws.add_image(img_excel, f'A{row}')
                
                # Set row height to accommodate image
                ws.row_dimensions[row].height = img_excel.height * 0.75
            except Exception as e:
                print(f"Error adding image {image_files[i]}: {e}")
                ws.cell(row=row, column=1).value = f"[Image: {os.path.basename(image_files[i])}]"
        
        # Add script if available
        slide_num = i + 1
        if slide_num in slides_dict:
            ws.cell(row=row, column=2).value = slides_dict[slide_num]
    
    # Save workbook
    if output_dir is None:
        output_dir = BASE_DIR / 'data'
    
    ppt_basename = os.path.basename(ppt_path).rsplit('.', 1)[0]
    output_file = os.path.join(output_dir, f"{ppt_basename}_{algo}_{model_provider}_lecture.xlsx")
    wb.save(output_file)
    print(f"Excel file created: {output_file}")

def main():
    parser = argparse.ArgumentParser(description='Convert lecture scripts and images to Excel')
    parser.add_argument('--ppt_path', '-p', help='Path to the PowerPoint file')
    parser.add_argument('--algo', '-a', help='Algorithm used for generation (e.g., vlm)')
    parser.add_argument('--model_provider', '-mp', help='Model provider (e.g., claude, gemini)')
    parser.add_argument('--output-dir', '-o', help='Directory to save the Excel file')
    
    args = parser.parse_args()
    scripts_to_excel(args.ppt_path, args.algo, args.model_provider, args.output_dir)

if __name__ == '__main__':
    main()
