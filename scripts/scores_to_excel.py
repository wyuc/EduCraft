import argparse
import json
import os
from pathlib import Path
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from config import BASE_DIR

# Color definition for highlighting cells
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Define metric groups for better organization
SLIDE_METRICS = [
    "avg_content_relevance",
    "avg_expressive_clarity", 
    "avg_logical_structure",
    "avg_audience_engagement",
    "avg_slide_score"
]

GLOBAL_METRICS = [
    "avg_narrative_flow",
    "avg_information_hierarchy",
    "avg_conceptual_integration",
    "avg_thematic_consistency", 
    "avg_cross_referencing",
    "avg_global_score"
]

SUMMARY_METRICS = [
    "avg_combined_score",
    "file_count"
]

# Friendly names for metrics
METRIC_NAMES = {
    "avg_content_relevance": "Content Relevance",
    "avg_expressive_clarity": "Expressive Clarity",
    "avg_logical_structure": "Logical Structure",
    "avg_audience_engagement": "Audience Engagement",
    "avg_slide_score": "Overall Slide Score",
    
    "avg_narrative_flow": "Narrative Flow",
    "avg_information_hierarchy": "Information Hierarchy",
    "avg_conceptual_integration": "Conceptual Integration",
    "avg_thematic_consistency": "Thematic Consistency",
    "avg_cross_referencing": "Cross Referencing",
    "avg_global_score": "Overall Global Score",
    
    "avg_combined_score": "Combined Score",
    "file_count": "Number of Files"
}

def load_scores_json(json_path):
    """
    Load scores from JSON file
    
    Args:
        json_path: Path to the JSON file
        
    Returns:
        dict: Loaded JSON data
    """
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)

def create_algorithm_worksheet(wb, algo_name, algo_data):
    """
    Create a worksheet for a specific algorithm with all its providers
    
    Args:
        wb: Workbook object
        algo_name: Algorithm name
        algo_data: Data for this algorithm (dict: provider -> metrics)
    """
    # Create worksheet
    ws = wb.create_sheet(title=algo_name[:31])  # Excel has 31 char limit for sheet names
    
    # Add title
    ws['A1'] = f"Performance Metrics for {algo_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Setup headers
    ws['A3'] = "Metric"
    ws.column_dimensions['A'].width = 25
    
    # Add provider headers
    providers = list(algo_data.keys())
    for i, provider in enumerate(providers):
        col = i + 2
        ws.cell(row=3, column=col, value=provider)
        ws.cell(row=3, column=col).font = Font(bold=True)
        # Set column width
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Add metrics by category
    current_row = 4
    
    # Slide metrics section
    ws.cell(row=current_row, column=1, value="SLIDE METRICS")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.merge_cells(f'A{current_row}:J{current_row}')
    current_row += 1
    
    for metric in SLIDE_METRICS:
        ws.cell(row=current_row, column=1, value=METRIC_NAMES.get(metric, metric))
        for i, provider in enumerate(providers):
            col = i + 2
            value = algo_data[provider].get(metric, "N/A")
            cell = ws.cell(row=current_row, column=col, value=value)
            if isinstance(value, (int, float)):
                cell.number_format = '0.00'
        current_row += 1
    
    current_row += 1
    
    # Global metrics section
    ws.cell(row=current_row, column=1, value="GLOBAL METRICS")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.merge_cells(f'A{current_row}:J{current_row}')
    current_row += 1
    
    for metric in GLOBAL_METRICS:
        ws.cell(row=current_row, column=1, value=METRIC_NAMES.get(metric, metric))
        for i, provider in enumerate(providers):
            col = i + 2
            value = algo_data[provider].get(metric, "N/A")
            cell = ws.cell(row=current_row, column=col, value=value)
            if isinstance(value, (int, float)):
                cell.number_format = '0.00'
        current_row += 1
    
    current_row += 1
    
    # Summary metrics section
    ws.cell(row=current_row, column=1, value="SUMMARY")
    ws.cell(row=current_row, column=1).font = Font(bold=True)
    ws.merge_cells(f'A{current_row}:J{current_row}')
    current_row += 1
    
    for metric in SUMMARY_METRICS:
        ws.cell(row=current_row, column=1, value=METRIC_NAMES.get(metric, metric))
        for i, provider in enumerate(providers):
            col = i + 2
            value = algo_data[provider].get(metric, "N/A")
            cell = ws.cell(row=current_row, column=col, value=value)
            if isinstance(value, (int, float)) and metric != "file_count":
                cell.number_format = '0.00'
        current_row += 1
    
    # Add conditional formatting to score cells
    for row in range(5, current_row):
        # Skip headers and file_count
        if ws.cell(row=row, column=1).value in ["SLIDE METRICS", "GLOBAL METRICS", "SUMMARY", "Number of Files"]:
            continue
            
        # Get range of cells for this row (excluding first column which is the metric name)
        data_range = f'B{row}:{get_column_letter(len(providers) + 1)}{row}'
        
        # Add color scale (red to green)
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            )
        )

def create_comparison_worksheet(wb, scores_data):
    """
    Create a comparison worksheet with top-level metrics across all algorithms
    
    Args:
        wb: Workbook object
        scores_data: Complete scores data
    """
    # Create worksheet
    ws = wb.create_sheet(title="Comparison")
    
    # Add title
    ws['A1'] = "Algorithm Performance Comparison"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Collect data for key metrics
    key_metrics = ['avg_slide_score', 'avg_global_score', 'avg_combined_score']
    comparison_data = {}
    
    # First, collect all algorithms and providers
    all_algorithms = []
    all_providers = set()
    
    for algo, providers in scores_data.items():
        all_algorithms.append(algo)
        for provider in providers:
            all_providers.add(provider)
    
    all_algorithms.sort()
    all_providers = sorted(list(all_providers))
    
    # Create headers
    ws['A3'] = "Algorithm"
    ws.column_dimensions['A'].width = 20
    
    for i, metric in enumerate(key_metrics):
        row = i + 3
        ws.cell(row=row, column=1, value=METRIC_NAMES.get(metric, metric))
        ws.cell(row=row, column=1).font = Font(bold=True)
    
    # Add provider columns
    for j, provider in enumerate(all_providers):
        col = j + 2
        ws.cell(row=2, column=col, value=provider)
        ws.cell(row=2, column=col).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Populate data
    row_offset = 3
    for i, algo in enumerate(all_algorithms):
        row = row_offset + i
        
        # Add algorithm name
        ws.cell(row=row, column=1, value=algo)
        
        # Add metric values for each provider
        for j, provider in enumerate(all_providers):
            col = j + 2
            
            # Check if the algorithm-provider combination exists
            if provider in scores_data.get(algo, {}):
                # Combined score
                value = scores_data[algo][provider].get('avg_combined_score', 'N/A')
                cell = ws.cell(row=row, column=col, value=value)
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00'
            else:
                ws.cell(row=row, column=col, value='N/A')
    
    # Add conditional formatting for the scores
    data_range = f'B{row_offset}:{get_column_letter(len(all_providers) + 1)}{row_offset + len(all_algorithms) - 1}'
    
    ws.conditional_formatting.add(
        data_range,
        ColorScaleRule(
            start_type='min', start_color='F8696B',
            mid_type='percentile', mid_value=50, mid_color='FFEB84',
            end_type='max', end_color='63BE7B'
        )
    )
    
    # Add a detailed breakdown section
    row_offset = row_offset + len(all_algorithms) + 2
    
    ws.cell(row=row_offset, column=1, value="Detailed Metric Breakdown")
    ws.cell(row=row_offset, column=1).font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row_offset}:J{row_offset}')
    ws.cell(row=row_offset, column=1).alignment = Alignment(horizontal='center')
    
    row_offset += 2
    
    # Add each metric in separate sections
    for metric in SLIDE_METRICS + GLOBAL_METRICS:
        # Skip the overall scores as they're in the top section
        if metric in key_metrics:
            continue
            
        ws.cell(row=row_offset, column=1, value=METRIC_NAMES.get(metric, metric))
        ws.cell(row=row_offset, column=1).font = Font(bold=True)
        
        # Add provider headers for this metric
        for j, provider in enumerate(all_providers):
            col = j + 2
            ws.cell(row=row_offset, column=col, value=provider)
            ws.cell(row=row_offset, column=col).font = Font(bold=True)
        
        row_offset += 1
        
        # Add data for each algorithm
        for i, algo in enumerate(all_algorithms):
            row = row_offset + i
            
            # Add algorithm name
            ws.cell(row=row, column=1, value=algo)
            
            # Add metric values for each provider
            for j, provider in enumerate(all_providers):
                col = j + 2
                
                # Check if the algorithm-provider combination exists
                if provider in scores_data.get(algo, {}):
                    value = scores_data[algo][provider].get(metric, 'N/A')
                    cell = ws.cell(row=row, column=col, value=value)
                    if isinstance(value, (int, float)):
                        cell.number_format = '0.00'
                else:
                    ws.cell(row=row, column=col, value='N/A')
        
        # Add conditional formatting for this metric
        metric_data_range = f'B{row_offset}:{get_column_letter(len(all_providers) + 1)}{row_offset + len(all_algorithms) - 1}'
        
        ws.conditional_formatting.add(
            metric_data_range,
            ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            )
        )
        
        row_offset += len(all_algorithms) + 2

def export_scores_to_excel(scores_data, output_path):
    """
    Export scores data to Excel file
    
    Args:
        scores_data: Scores data (dict: algorithm -> provider -> metrics)
        output_path: Path to save the Excel file
        
    Returns:
        str: Path to the created Excel file
    """
    # Create workbook
    wb = Workbook()
    
    # Remove default worksheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    # Create comparison sheet (overview)
    create_comparison_worksheet(wb, scores_data)
    
    # Create individual sheets for each algorithm
    for algo, algo_data in scores_data.items():
        create_algorithm_worksheet(wb, algo, algo_data)
    
    # Save workbook
    wb.save(output_path)
    print(f"Exported scores to: {output_path}")
    
    return output_path

def main():
    parser = argparse.ArgumentParser(description='Convert evaluation scores JSON to Excel')
    parser.add_argument('--input', '-i', default='data/eval_results/scores_report_test_set_lecgen.json',
                       help='Path to the input JSON file')
    parser.add_argument('--output', '-o', help='Path to save the Excel file')
    
    args = parser.parse_args()
    
    # Check if input file exists
    if not os.path.exists(args.input):
        print(f"Error: Input file {args.input} does not exist")
        return
    
    # Load scores
    print(f"Loading scores from {args.input}")
    scores_data = load_scores_json(args.input)
    
    # Determine output path
    if args.output:
        output_path = args.output
    else:
        input_path = Path(args.input)
        output_dir = input_path.parent
        output_path = output_dir / f"{input_path.stem}.xlsx"
    
    # Make sure output directory exists
    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    
    # Export to Excel
    export_scores_to_excel(scores_data, output_path)

if __name__ == '__main__':
    main() 