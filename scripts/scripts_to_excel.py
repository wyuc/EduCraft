import os
import json
import argparse
import glob
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from PIL import Image as PILImage
from storage import ScriptStorage
from utils import get_images_dir
from config import BASE_DIR

def natural_sort_key(s):
    """
    Natural sort key function for sorting filenames with numbers.
    This ensures that e.g., "2.png" comes before "10.png"
    """
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', s)]

def get_all_completed_tasks_for_file(input_path):
    """
    Get all completed tasks for a given file across all algorithms and providers.
    
    Args:
        input_path (str): Path to the input file
        
    Returns:
        list: List of tuples (algo, provider) for which completed tasks exist
    """
    storage = ScriptStorage()
    
    input_path = Path(input_path)
    file_name = input_path.stem
    
    # Get all tasks from storage
    all_tasks = storage.get_all_tasks()
    
    # Filter for completed tasks matching the file name
    completed_tasks = [
        task for task in all_tasks 
        if task['file_name'] == file_name 
        and task['status'] == 'completed'
        and task['algo'] is not None
    ]
    
    # Extract unique algo-provider combinations
    algo_provider_pairs = set()
    for task in completed_tasks:
        algo_provider_pairs.add((task['algo'], task['model_provider']))
    
    return sorted(list(algo_provider_pairs))

def filter_algorithms(algo_provider_pairs, selected_algos, selected_providers=None):
    """
    Filter algorithm-provider pairs to only include selected algorithms and/or providers.
    
    Args:
        algo_provider_pairs: List of (algo, provider) tuples
        selected_algos: List of algorithm names to include
        selected_providers: List of provider names to include
        
    Returns:
        list: Filtered list of (algo, provider) tuples
    """
    # If neither algos nor providers specified, return all
    if not selected_algos and not selected_providers:
        return algo_provider_pairs
    
    filtered_pairs = algo_provider_pairs
    
    # Filter by algorithms if specified
    if selected_algos:
        # Convert to lowercase for case-insensitive comparison
        selected_algos_lower = [algo.lower() for algo in selected_algos]
        
        # Filter pairs where the algorithm is in the selected list
        filtered_pairs = [
            (algo, provider) for (algo, provider) in filtered_pairs
            if algo.lower() in selected_algos_lower
        ]
    
    # Filter by providers if specified
    if selected_providers:
        # Convert to lowercase for case-insensitive comparison
        selected_providers_lower = [provider.lower() for provider in selected_providers]
        
        # Filter pairs where the provider is in the selected list
        filtered_pairs = [
            (algo, provider) for (algo, provider) in filtered_pairs
            if provider.lower() in selected_providers_lower
        ]
    
    return filtered_pairs

def get_slides_for_task(task_id):
    """
    Get slides content for a specific task.
    
    Args:
        task_id: Task ID
    
    Returns:
        dict: Dictionary with slide numbers as keys and content as values
    """
    storage = ScriptStorage()
    
    # Get all slides for this task
    slides = storage.get_all_slides(task_id)
    if not slides:
        print(f"Error: No slides found for task {task_id}")
        return {}
    
    # Convert slides to dict for easier access by slide number
    slides_dict = {slide['slide_num']: slide['content'] for slide in slides}
    return slides_dict

def get_task_id(file_name, algo, provider):
    """
    Get the task ID for a specific file, algorithm, and provider.
    
    Args:
        file_name: Name of the file
        algo: Algorithm name
        provider: Model provider
    
    Returns:
        str: Task ID or None if not found
    """
    storage = ScriptStorage()
    
    # Get the latest completed task matching the criteria
    latest_task = storage.get_latest_completed_task(file_name, algo, provider)
    
    if not latest_task:
        print(f"Error: No completed tasks found for {file_name} with algorithm {algo} and provider {provider}")
        return None
    
    return latest_task['task_id']

def load_slide_scores(file_name, eval_file_path=None):
    """
    Load slide-level evaluation scores from the evaluation results JSON file.
    
    Args:
        file_name: Name of the file (without extension)
        eval_file_path: Optional path to a specific evaluation results file
    
    Returns:
        dict: Dictionary with (algo, provider, slide_num) as keys and 
              a dict containing individual metric scores and the average score as values.
              Example: {(algo, provider, 1): {'content_relevance': 4, 'expressive_clarity': 5, ..., 'average': 4.5}}
    """
    # If a specific evaluation file is provided, use it
    if eval_file_path and os.path.exists(eval_file_path):
        eval_file = eval_file_path
    else:
        # Otherwise try to locate the default evaluation results file
        eval_dir = BASE_DIR / 'data' / 'eval_results'
        eval_file = eval_dir / f"evaluation_results_{file_name}_lecgen.json"
        
        # Also check for files generated by eval/main.py
        if not os.path.exists(eval_file):
            # Look for evaluation_results_{file_name}_*.json files
            potential_files = list(eval_dir.glob(f"evaluation_results_{file_name}_*.json"))
            if potential_files:
                # Use the most recent file
                eval_file = max(potential_files, key=os.path.getmtime)
    
    if not os.path.exists(eval_file):
        print(f"Warning: Evaluation results file not found at {eval_file}")
        return {}
    
    try:
        with open(eval_file, 'r', encoding='utf-8') as f:
            eval_data = json.load(f)
        print(f"Loading scores from: {eval_file}")
    except Exception as e:
        print(f"Error loading evaluation results: {e}")
        return {}
    
    # Extract slide-level scores
    slide_scores = {}
    score_count = 0
    
    # Metric keys in order they appear in the scores array or as keys
    metric_keys = [
        'content_relevance',
        'expressive_clarity',
        'logical_structure',
        'audience_engagement'
    ]
    
    for algo, algo_data in eval_data.items():
        for provider, provider_data in algo_data.items():
            # Check all possible file names in the provider data
            files_to_check = []
            
            # Add the exact file name
            files_to_check.append(file_name)
            
            # Also check for any files that contain the file_name as a substring
            # This helps with cases where the file might be stored under a slightly different name
            for key in provider_data.keys():
                if isinstance(key, str) and file_name in key:
                    files_to_check.append(key)
            
            # Process each potential file
            for file_key in files_to_check:
                if file_key in provider_data and 'slide_evaluation' in provider_data[file_key]:
                    slide_eval = provider_data[file_key]['slide_evaluation']
                    
                    # Format 1: Using per_slide_scores array
                    if 'per_slide_scores' in slide_eval:
                        for slide_data in slide_eval['per_slide_scores']:
                            slide_num = slide_data.get('slide')
                            scores_array = slide_data.get('scores', [])
                            
                            if slide_num is not None and len(scores_array) == len(metric_keys):
                                current_scores = {}
                                valid_scores = []
                                for i, metric in enumerate(metric_keys):
                                    score_val = scores_array[i]
                                    current_scores[metric] = score_val
                                    if isinstance(score_val, (int, float)):
                                        valid_scores.append(score_val)
                                
                                # Calculate average score
                                if valid_scores:
                                    avg_score = sum(valid_scores) / len(valid_scores)
                                    current_scores['average'] = avg_score
                                else:
                                    current_scores['average'] = None # Or handle as needed
                                    
                                slide_scores[(algo, provider, slide_num)] = current_scores
                                score_count += 1
                    
                    # Format 2: Using slides object with metric values
                    elif 'slides' in slide_eval:
                        for slide_num_str, slide_data in slide_eval['slides'].items():
                            try:
                                slide_num = int(slide_num_str)
                                current_scores = {}
                                valid_scores = []
                                
                                for metric in metric_keys:
                                    if metric in slide_data:
                                        score_val = slide_data[metric]
                                        current_scores[metric] = score_val
                                        if isinstance(score_val, (int, float)):
                                            valid_scores.append(score_val)
                                
                                if valid_scores: # Only add if we found some scores
                                    avg_score = sum(valid_scores) / len(valid_scores)
                                    current_scores['average'] = avg_score
                                    slide_scores[(algo, provider, slide_num)] = current_scores
                                    score_count += 1
                                else:
                                    # Store anyway if you want to capture slides even without valid scores
                                    # current_scores['average'] = None
                                    # slide_scores[(algo, provider, slide_num)] = current_scores
                                    pass # Or skip this slide if no scores found

                            except ValueError:
                                continue
    
    print(f"Loaded detailed scores for {score_count} slides across {len(slide_scores)} algorithm-provider-slide combinations")
    return slide_scores

def create_side_by_side_excel(input_path, output_dir=None, selected_algos=None, selected_providers=None, eval_file_path=None):
    """
    Create a single Excel file with all algorithms in separate columns.
    
    Args:
        input_path: Path to the input file
        output_dir: Directory to save the Excel file
        selected_algos: List of algorithm names to include (None for all)
        selected_providers: List of provider names to include (None for all)
        eval_file_path: Optional path to a specific evaluation results file
    
    Returns:
        str: Path to the created Excel file or None if failed
    """
    input_path = Path(input_path)
    file_name = input_path.stem
    
    # Get all algorithms and providers for this file
    algo_provider_pairs = get_all_completed_tasks_for_file(input_path)
    
    if not algo_provider_pairs:
        print(f"No completed tasks found for file: {file_name}")
        return None
    
    # Filter algorithms and providers if specified
    if selected_algos or selected_providers:
        original_count = len(algo_provider_pairs)
        algo_provider_pairs = filter_algorithms(algo_provider_pairs, selected_algos, selected_providers)
        filtered_count = len(algo_provider_pairs)
        
        if filtered_count == 0:
            available_algos = set(algo for algo, _ in get_all_completed_tasks_for_file(input_path))
            available_providers = set(provider for _, provider in get_all_completed_tasks_for_file(input_path))
            
            if selected_algos:
                print(f"No matching algorithms found among the specified algorithms: {', '.join(selected_algos)}")
                print(f"Available algorithms: {', '.join(available_algos)}")
            
            if selected_providers:
                print(f"No matching providers found among the specified providers: {', '.join(selected_providers)}")
                print(f"Available providers: {', '.join(available_providers)}")
                
            return None
        
        print(f"Selected {filtered_count} of {original_count} available algorithm-provider combinations")
    
    # Load slide scores (now returns detailed dict)
    slide_scores = load_slide_scores(file_name, eval_file_path)
    has_scores = len(slide_scores) > 0
    if has_scores:
        print(f"Loaded detailed evaluation scores for {len(slide_scores)} slides") # Updated print message slightly
    else:
        print("No evaluation scores found, will display scripts only")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Lecture Comparison"
    
    # Get images directory
    images_dir = get_images_dir(input_path)
    
    # Find image files if available
    image_files = []
    if images_dir and os.path.exists(images_dir):
        image_files = glob.glob(os.path.join(images_dir, '*.*'))
        image_extensions = ['.png', '.jpg', '.jpeg', '.gif', '.bmp']
        image_files = [f for f in image_files if os.path.splitext(f)[1].lower() in image_extensions]
        
        # Use natural sorting to ensure correct numerical order
        image_files.sort(key=natural_sort_key)
    
    # Collect all slides from all algorithms
    all_slides = {}
    task_mapping = {}
    
    for algo, provider in algo_provider_pairs:
        task_id = get_task_id(file_name, algo, provider)
        if task_id:
            slides_dict = get_slides_for_task(task_id)
            if slides_dict:
                all_slides[(algo, provider)] = slides_dict
                task_mapping[(algo, provider)] = task_id
    
    if not all_slides:
        print("No slides found for any algorithm")
        return None
    
    # Determine all slide numbers across all algorithms
    all_slide_numbers = set()
    for slides_dict in all_slides.values():
        all_slide_numbers.update(slides_dict.keys())
    
    max_slides = max(all_slide_numbers) if all_slide_numbers else 0
    
    # Set up column headers
    ws.cell(row=1, column=1).value = "Slide"
    
    column_index = 2
    column_mapping = {}
    
    for algo, provider in algo_provider_pairs:
        if (algo, provider) in all_slides:
            # Add headers for script and score columns
            script_col = column_index
            score_col = column_index + 1
            
            ws.cell(row=1, column=script_col).value = f"{algo}-{provider}"
            if has_scores:
                # Update score header
                ws.cell(row=1, column=score_col).value = f"Scores (Avg/R/C/S/E)" 
            
            column_mapping[(algo, provider)] = (script_col, score_col)
            
            column_index += 2 if has_scores else 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 75  # Slide column
    
    for algo, provider in algo_provider_pairs:
        if (algo, provider) in column_mapping:
            script_col, score_col = column_mapping[(algo, provider)]
            
            # Set script column width
            script_col_letter = get_column_letter(script_col)
            ws.column_dimensions[script_col_letter].width = 70
            
            if has_scores:
                # Set score column width (increased)
                score_col_letter = get_column_letter(score_col)
                ws.column_dimensions[score_col_letter].width = 20 # Increased width

    # Helper function for formatting scores
    def format_score(s, precision=1):
        if isinstance(s, (int, float)):
            return f"{s:.{precision}f}"
        return str(s) if s is not None else 'N/A'

    # Add images and scripts row by row
    for slide_num in range(1, max_slides + 1):
        row = slide_num + 1  # Start from row 2 (after headers)
        
        # Add image if available
        if image_files and slide_num <= len(image_files):
            try:
                # Get image dimensions
                img = PILImage.open(image_files[slide_num - 1])
                width, height = img.size
                
                # Resize image for Excel
                max_height = 400
                max_width = 450
                scale = min(max_width/width, max_height/height)
                
                img_excel = Image(image_files[slide_num - 1])
                img_excel.width = width * scale
                img_excel.height = height * scale
                
                # Add to cell
                ws.add_image(img_excel, f'A{row}')
                
                # Set row height to accommodate image
                ws.row_dimensions[row].height = img_excel.height * 0.75
            except Exception as e:
                print(f"Error adding image for slide {slide_num}: {e}")
                ws.cell(row=row, column=1).value = f"[Slide {slide_num}]"
        else:
            ws.cell(row=row, column=1).value = f"[Slide {slide_num}]"
        
        # Add scripts and scores for each algorithm
        for (algo, provider), (script_col, score_col) in column_mapping.items():
            # Add script content
            slides_dict = all_slides.get((algo, provider), {})
            if slide_num in slides_dict:
                script_cell = ws.cell(row=row, column=script_col, value=slides_dict[slide_num])
                script_cell.alignment = Alignment(wrap_text=True, vertical='top') # Ensure script wrap
            
            # Add score if available
            if has_scores:
                score_data = slide_scores.get((algo, provider, slide_num), None)
                if score_data is not None:
                    # Extract scores
                    avg = score_data.get('average')
                    rel = score_data.get('content_relevance')
                    clar = score_data.get('expressive_clarity')
                    struc = score_data.get('logical_structure')
                    eng = score_data.get('audience_engagement')

                    # Format cell text with line breaks
                    cell_text = (
                        f"Avg: {format_score(avg, 2)}\n"
                        f"R: {format_score(rel)} C: {format_score(clar)}\n"
                        f"S: {format_score(struc)} E: {format_score(eng)}"
                    )
                    
                    score_cell = ws.cell(row=row, column=score_col, value=cell_text)
                    # Apply word wrap and vertical alignment
                    score_cell.alignment = Alignment(wrap_text=True, vertical='top') 
                    
                    # Apply conditional formatting based on the average score
                    if isinstance(avg, (int, float)):
                        # Green for high scores (>= 4.0)
                        if avg >= 4.0:
                            score_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                            score_cell.font = Font(color="006100")
                        # Yellow for medium scores (>= 3.0)
                        elif avg >= 3.0:
                            score_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                            score_cell.font = Font(color="9C5700")
                        # Red for low scores (< 3.0)
                        else:
                            score_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                            score_cell.font = Font(color="9C0006")
                else:
                     # Handle case where score_data is missing for this slide/algo/provider
                     ws.cell(row=row, column=score_col, value="N/A")

    # Prepare output path
    if output_dir is None:
        output_dir = BASE_DIR / 'data'
    else:
        output_dir = Path(output_dir)
    
    # Create output directory if it doesn't exist
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # Create a meaningful filename that includes which algorithms are included
    filename_parts = []
    
    if selected_algos:
        algo_part = "_".join(selected_algos[:2])  # Include up to 2 algorithm names in the filename
        if len(selected_algos) > 2:
            algo_part += "_etc"
        filename_parts.append(algo_part)
    
    if selected_providers:
        provider_part = "_".join(selected_providers[:2])  # Include up to 2 provider names in the filename
        if len(selected_providers) > 2:
            provider_part += "_etc"
        filename_parts.append(provider_part)
    
    if filename_parts:
        output_file = os.path.join(output_dir, f"{file_name}_{'_'.join(filename_parts)}.xlsx")
    else:
        output_file = os.path.join(output_dir, f"{file_name}_side_by_side.xlsx")
    
    wb.save(output_file)
    print(f"Excel file created with side-by-side comparison: {output_file}")
    
    return output_file

def scripts_to_excel(input_path, algo=None, model_provider=None, output_dir=None, eval_file_path=None):
    """
    Create an Excel file with scripts from a PowerPoint file.
    
    Args:
        input_path (str): Path to the input file
        algo (str, optional): Algorithm name to filter by
        model_provider (str, optional): Model provider to filter by
        output_dir (str, optional): Directory to save the Excel file
        eval_file_path (str, optional): Path to the evaluation results file
    
    Returns:
        str: Path to the created Excel file or None if creation failed
    """
    selected_algos = [algo] if algo else None
    selected_providers = [model_provider] if model_provider else None
    
    return create_side_by_side_excel(
        input_path=input_path,
        output_dir=output_dir,
        selected_algos=selected_algos,
        selected_providers=selected_providers,
        eval_file_path=eval_file_path
    )

def list_available_algorithms(input_path):
    """
    List all available algorithms and providers for a given file.
    
    Args:
        input_path: Path to the input file
    """
    algo_provider_pairs = get_all_completed_tasks_for_file(input_path)
    
    if not algo_provider_pairs:
        print(f"No completed tasks found for file: {Path(input_path).stem}")
        return
    
    # Extract unique algorithms and providers
    unique_algos = sorted(set(algo for algo, _ in algo_provider_pairs))
    unique_providers = sorted(set(provider for _, provider in algo_provider_pairs))
    
    print("\nAvailable algorithms:")
    for algo in unique_algos:
        providers = sorted(provider for a, provider in algo_provider_pairs if a == algo)
        print(f"  {algo} (providers: {', '.join(providers)})")
    
    print("\nAvailable providers:")
    for provider in unique_providers:
        algos = sorted(algo for algo, p in algo_provider_pairs if p == provider)
        print(f"  {provider} (algorithms: {', '.join(algos)})")
    
    print("\nAvailable algorithm-provider combinations:")
    for algo, provider in sorted(algo_provider_pairs):
        print(f"  {algo}-{provider}")

def main():
    parser = argparse.ArgumentParser(description='Convert lecture scripts and images to Excel')
    parser.add_argument('input_path', help='Path to the input file')
    parser.add_argument('--output-dir', '-o', help='Directory to save the Excel file')
    parser.add_argument('--algos', '-a', nargs='+', help='Specific algorithms to include (e.g., vlm vlm_no_narrative)')
    parser.add_argument('--model-providers', '-mp', nargs='+', help='Specific model providers to include (e.g., gpt-4 claude)')
    parser.add_argument('--list', '-l', action='store_true', help='List available algorithms and providers for the input file')
    parser.add_argument('--eval-file', '-e', help='Path to a specific evaluation results file to read scores from')
    
    args = parser.parse_args()
    
    # List available algorithms and providers if requested
    if args.list:
        list_available_algorithms(args.input_path)
        return
    
    # Create a side-by-side comparison
    output_file = create_side_by_side_excel(
        args.input_path, 
        args.output_dir, 
        args.algos, 
        args.model_providers,
        args.eval_file
    )
    if output_file:
        print("Excel file creation completed successfully.")
    else:
        print("Failed to create Excel file.")

if __name__ == '__main__':
    main()
